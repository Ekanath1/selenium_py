import openpyxl
from openpyxl import Workbook
import pandas as pd

def read_data():
    wb = openpyxl.load_workbook(r'C:\\Users\\AKASH\\PycharmProjects\\sel\\Data\\Book1.xlsx')

    sheet = wb["Sheet1"]  # Make sure this sheet exists
    for row in sheet.iter_rows(min_row=2, values_only=True):
        print(row)
    table_data = [
        ["Name", "Age", "City","12", "23"],
        ["Alice", 30, "Delhi", "23", "ss"],
        ["Bob", 25, "Mumbai", "d3", "33"],
        ["Charlie", 28, "Bangalore", "sd", "3d"]
    ]


    wb = Workbook()
    ws = wb.active

    for row in table_data:
        ws.append(row)

    wb.save("Books.xlsx")
    print("Data written to Excel successfully.")

# read_data()

def read_data1():
    workbook = openpyxl.load_workbook(r'Book.xlsx')
    sheet_names = workbook.sheetnames

    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            print(row)

read_data1()
print("To read one sheet data")

def one_sheet_data():
    wb = openpyxl.load_workbook(r'File_Name.xlsx')
    sheet_name = wb['Sheet']
    for row in sheet_name.iter_rows(min_row=2, values_only=True):
        print(row)

# one_sheet_data()

def one_sheet():
    wb = openpyxl.load_workbook(r'Book.xlsx')
    sheet = wb['Sheet']
    for row in sheet.iter_rows(min_row=2, values_only=True):
        print(row)

one_sheet()